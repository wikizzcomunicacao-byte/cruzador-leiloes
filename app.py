from datetime import datetime
import io
import re
from urllib.parse import quote
from fpdf import FPDF
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import plotly.express as px
import streamlit as st

# ---------------------------------------------------------
# 1. CONFIGURAÇÃO DA PÁGINA
# ---------------------------------------------------------
st.set_page_config(
    page_title="Cruzador de Leilões Enterprise",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------
# 2. CREDENCIAIS DE ACESSO (ADMIN + TESTE)
# ---------------------------------------------------------
USUARIOS = {
    "administrador": "22029804",
    "teste": "teste123",
}

if "autenticado" not in st.session_state:
  st.session_state["autenticado"] = False
if "usuario_logado" not in st.session_state:
  st.session_state["usuario_logado"] = ""


# ---------------------------------------------------------
# 3. TELA DE LOGIN
# ---------------------------------------------------------
def tela_login():
  st.markdown("<br><br><br>", unsafe_allow_html=True)
  col1, col2, col3 = st.columns([1, 1.4, 1])

  with col2:
    st.markdown(
        """
        <div style="background-color: #FFFFFF; padding: 2.5rem; border-radius: 12px; border: 1px solid #E2E8F0; box-shadow: 0 4px 16px rgba(0,0,0,0.08);">
            <h2 style="text-align: center; color: #0052CC; margin-bottom: 0;">🏢 Cruzador Pro</h2>
            <p style="text-align: center; color: #64748B; font-size: 0.9rem;">Plataforma de Oportunidades em Leilões</p>
            <hr style="border: 0.5px solid #E2E8F0; margin-bottom: 1.5rem; margin-top: 1rem;">
        """,
        unsafe_allow_html=True,
    )

    user_input = st.text_input("👤 Usuário", key="login_user")
    pass_input = st.text_input("🔑 Senha", type="password", key="login_pass")

    st.write(" ")
    if st.button("🔓 Entrar no Sistema", use_container_width=True):
      if user_input in USUARIOS and USUARIOS[user_input] == pass_input:
        st.session_state["autenticado"] = True
        st.session_state["usuario_logado"] = user_input
        st.rerun()
      else:
        st.error("Usuário ou senha incorretos!")

    st.markdown("</div>", unsafe_allow_html=True)


# ---------------------------------------------------------
# 4. CONTROLE DE FLUXO (LOGIN vs APLICATIVO)
# ---------------------------------------------------------
if not st.session_state["autenticado"]:
  tela_login()
else:
  is_tester = st.session_state["usuario_logado"] == "teste"

  # BARRA LATERAL
  with st.sidebar:
    st.title("🏢 Cruzador Pro")
    if is_tester:
      st.info("⚠️ Modo de Teste: Visualização restrita.")
    else:
      st.caption(f"👤 Conectado: **{st.session_state['usuario_logado']}**")

    if st.button("🚪 Sair / Logout"):
      st.session_state["autenticado"] = False
      st.session_state["usuario_logado"] = ""
      if "df_final" in st.session_state:
        del st.session_state["df_final"]
      if "imoveis_selecionados" in st.session_state:
        del st.session_state["imoveis_selecionados"]
      st.rerun()

    st.divider()
    st.markdown("💡 *Dica: Use as abas para navegar entre os recursos.*")

  # ESTILIZAÇÃO CSS CUSTOMIZADA
  st.markdown(
      """
        <style>
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}
        
        .main .block-container {
            padding-top: 1.5rem;
            padding-bottom: 3rem;
        }

        .stTextInput input, .stSelectbox select, .stNumberInput input {
            color: #1E293B !important;
            opacity: 1 !important;
        }
        
        div.stButton > button:first-child {
            background-color: #0052CC;
            color: white;
            font-weight: bold;
            border-radius: 8px;
            padding: 0.6rem 1rem;
            border: none;
            width: 100%;
            transition: all 0.3s ease;
        }
        div.stButton > button:first-child:hover {
            background-color: #003D99;
            color: white;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
        }
        
        div[data-testid="stMetric"] {
            background-color: #FFFFFF;
            border: 1px solid #E2E8F0;
            padding: 1rem;
            border-radius: 12px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.04);
        }
        div[data-testid="stMetricValue"] {
            font-size: 1.6rem;
            font-weight: 700;
            color: #0052CC;
        }
        
        .property-card {
            background-color: #FFFFFF;
            border: 1px solid #E2E8F0;
            border-radius: 12px;
            padding: 1.2rem;
            margin-bottom: 1rem;
            box-shadow: 0 2px 8px rgba(0,0,0,0.05);
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        }
        .property-card:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 14px rgba(0,0,0,0.1);
        }
        .badge-type {
            background-color: #E0E7FF;
            color: #3730A3;
            padding: 4px 10px;
            border-radius: 20px;
            font-size: 0.82rem;
            font-weight: 600;
            display: inline-block;
            margin-bottom: 8px;
        }
        .price-main {
            font-size: 1.35rem;
            font-weight: bold;
            color: #166534;
        }
        .price-profit {
            font-size: 1.05rem;
            font-weight: 700;
            color: #2563EB;
        }
        .price-costs {
            font-size: 0.88rem;
            color: #64748B;
        }
        .price-old {
            font-size: 0.9rem;
            color: #9CA3AF;
            text-decoration: line-through;
        }
        </style>
    """,
      unsafe_allow_html=True,
  )

  # FUNÇÕES AUXILIARES DE TRATAMENTO
  def clean_ascii(text):
    if not isinstance(text, str):
      text = str(text) if text is not None else ""
    replacements = {
        "á": "a",
        "à": "a",
        "â": "a",
        "ã": "a",
        "ä": "a",
        "Á": "A",
        "À": "A",
        "Â": "A",
        "Ã": "A",
        "é": "e",
        "è": "e",
        "ê": "e",
        "ë": "e",
        "É": "E",
        "È": "E",
        "Ê": "E",
        "í": "i",
        "ì": "i",
        "î": "i",
        "ï": "i",
        "Í": "I",
        "Ì": "I",
        "Î": "I",
        "ó": "o",
        "ò": "o",
        "ô": "o",
        "õ": "o",
        "ö": "o",
        "Ó": "O",
        "Ò": "O",
        "Ô": "O",
        "Õ": "O",
        "ú": "u",
        "ù": "u",
        "û": "u",
        "ü": "u",
        "Ú": "U",
        "Ù": "U",
        "Û": "U",
        "ç": "c",
        "Ç": "C",
        "º": "o",
        "ª": "a",
        "²": "2",
    }
    for k, v in replacements.items():
      text = text.replace(k, v)
    return text.encode("latin-1", "replace").decode("latin-1")

  def normalize(text):
    if not isinstance(text, str):
      return ""
    text = text.lower()
    text = re.sub(r"[áàâãä]", "a", text)
    text = re.sub(r"[éèêë]", "e", text)
    text = re.sub(r"[íìîï]", "i", text)
    text = re.sub(r"[óòôõö]", "o", text)
    text = re.sub(r"[úùûü]", "u", text)
    text = re.sub(r"[ç]", "c", text)
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return " ".join(text.split())

  def parse_budget(budget_str):
    budget_str = str(budget_str)
    if "100.000" in budget_str:
      return 0, 100000
    elif "200.000" in budget_str:
      return 0, 200000
    elif "300.000" in budget_str:
      return 0, 300000
    elif "400.000" in budget_str:
      return 0, 400000
    elif "500.000" in budget_str:
      return 500000, 999999999
    else:
      return 0, 999999999

  def parse_types(tipos_str):
    norm_t = normalize(tipos_str)
    res = set()
    if "apartamento" in norm_t:
      res.add("Apartamento")
    if "casa" in norm_t:
      res.add("Casa")
    if "terreno" in norm_t:
      res.add("Terreno")
    if "lote comercial" in norm_t or "comercial" in norm_t:
      res.add("Comercial")
    if "outros" in norm_t:
      res.update(["Area Rural", "Comercial", "Indefinido", "Vaga de Garagem"])
    return list(res)

  class InformativoLeiloesPDF(FPDF):

    def header(self):
      if self.page_no() > 1:
        self.set_font("Arial", "B", 8)
        self.set_text_color(100, 116, 139)
        self.cell(
            0,
            5,
            clean_ascii("LOURENCO COLOMBO E ROZANI - ADVOCACIA E LEILOES"),
            0,
            1,
            "L",
        )
        self.set_draw_color(226, 232, 240)
        self.line(10, 12, 200, 12)
        self.ln(3)

    def footer(self):
      self.set_y(-12)
      self.set_font("Arial", "I", 8)
      self.set_text_color(128, 128, 128)
      self.cell(
          0,
          10,
          clean_ascii(
              f"Informativo de Imoveis em Leilao - Pagina {self.page_no()}"
          ),
          0,
          0,
          "C",
      )

  def gerar_pdf_informativo(nome_investidor, df_inv):
    pdf = InformativoLeiloesPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.add_page()
    pdf.ln(15)

    pdf.set_font("Arial", "B", 22)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(0, 10, clean_ascii("INFORMATIVO DE IMÓVEIS"), 0, 1, "C")
    pdf.cell(0, 10, clean_ascii("EM LEILÃO"), 0, 1, "C")
    pdf.ln(8)

    pdf.set_font("Arial", "B", 11)
    pdf.set_text_color(70, 80, 95)
    pdf.cell(0, 6, clean_ascii("LOURENÇO COLOMBO E ROZANI"), 0, 1, "C")
    pdf.set_font("Arial", "", 8.5)
    pdf.cell(0, 5, clean_ascii("ADVOCACIA E LEILÕES IMOBILIÁRIOS"), 0, 1, "C")
    pdf.ln(15)

    pdf.set_font("Arial", "", 10.5)
    pdf.set_text_color(51, 65, 85)
    pdf.multi_cell(
        0,
        6,
        clean_ascii(
            f"Prezado(a) Investidor(a) {nome_investidor}.\n\n"
            "É com prazer que apresentamos nosso informativo de imóveis "
            "disponíveis em leilão. Esta é uma oportunidade única para adquirir "
            "bens com preços atrativos e grande potencial de valorização!\n\n"
            "Confira adiante algumas das propriedades em destaque, de "
            "acordo com as regiões selecionadas por você."
        ),
    )
    pdf.ln(15)

    cid_req = (
        str(df_inv["Cidades Solicitadas"].iloc[0])
        if not df_inv.empty
        else "N/A"
    )
    faixa_req = (
        str(df_inv["Faixa Solicitada"].iloc[0]) if not df_inv.empty else "N/A"
    )

    pdf.set_fill_color(248, 250, 252)
    pdf.set_draw_color(226, 232, 240)
    pdf.rect(15, 145, 180, 28, style="DF")
    pdf.set_xy(20, 150)
    pdf.set_font("Arial", "B", 9.5)
    pdf.cell(0, 5, clean_ascii("PARÂMETROS DO INVESTIDOR:"), 0, 1)
    pdf.set_x(20)
    pdf.set_font("Arial", "", 9)
    pdf.cell(
        0, 5, clean_ascii(f"Regiões Selecionadas: {cid_req}"), 0, 1
    )
    pdf.set_x(20)
    pdf.cell(
        0, 5, clean_ascii(f"Faixa de Orçamento: {faixa_req}"), 0, 1
    )

    if df_inv.empty:
      pdf.add_page()
      pdf.set_font("Arial", "B", 11)
      pdf.set_text_color(200, 0, 0)
      pdf.cell(
          0,
          10,
          clean_ascii(
              "NÃO EXISTEM OPORTUNIDADES VIÁVEIS NAS REGIÕES SELECIONADAS"
          ),
          0,
          1,
          "C",
      )
    else:
      tipos_unicos = df_inv["Tipo de Bem"].unique()

      for tipo in tipos_unicos:
        pdf.add_page()
        pdf.set_font("Arial", "B", 13)
        pdf.set_text_color(30, 41, 59)
        pdf.cell(0, 8, clean_ascii(str(tipo).upper()), 0, 1, "L")
        pdf.set_draw_color(100, 116, 139)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(4)

        df_tipo = df_inv[df_inv["Tipo de Bem"] == tipo]

        for idx, (_, row) in enumerate(df_tipo.iterrows(), start=1):
          if pdf.get_y() > 235:
            pdf.add_page()

          y_inicial = pdf.get_y()

          pdf.set_fill_color(252, 253, 255)
          pdf.set_draw_color(210, 215, 225)
          pdf.rect(10, y_inicial, 190, 48, style="DF")

          pdf.set_xy(12, y_inicial + 2)
          pdf.set_font("Arial", "B", 9.5)
          pdf.set_text_color(0, 82, 204)
          titulo_num = f"{idx:02d}. {str(row['Título do Imóvel']).upper()}"
          pdf.cell(186, 5, clean_ascii(titulo_num[:85]), 0, 1)

          pdf.set_font("Arial", "", 8)
          pdf.set_text_color(50, 60, 75)

          col1_x = 14
          col2_x = 75
          col3_x = 135
          current_y = pdf.get_y() + 1

          pdf.set_xy(col1_x, current_y)
          pdf.cell(
              60, 4, clean_ascii(f"Matrícula: {row.get('Matrícula', 'N/I')}"), 0, 0
          )
          pdf.set_xy(col2_x, current_y)
          pdf.cell(
              60,
              4,
              clean_ascii(
                  f"Valor Avaliação: R$ {row['Valor de Avaliação (R$)']:,.2f}"
              ),
              0,
              0,
          )
          pdf.set_xy(col3_x, current_y)
          pdf.cell(
              60,
              4,
              clean_ascii(
                  f"Instituição: {row.get('Instituição', 'Judicial/Caixa')}"
              ),
              0,
              1,
          )

          current_y += 4.5
          pdf.set_xy(col1_x, current_y)
          pdf.cell(
              60,
              4,
              clean_ascii(
                  f"Área do Terreno: {row.get('Área do Terreno', 'N/I')} m²"
              ),
              0,
              0,
          )
          pdf.set_xy(col2_x, current_y)
          pdf.cell(
              60,
              4,
              clean_ascii(
                  f"Lance Mínimo: R$ {row['Preço do Leilão (R$)']:,.2f}"
              ),
              0,
              0,
          )
          pdf.set_xy(col3_x, current_y)
          pdf.cell(
              60,
              4,
              clean_ascii(
                  f"Modalidade: {row.get('Modalidade', 'Leilão')}"
              ),
              0,
              1,
          )

          current_y += 4.5
          pdf.set_xy(col1_x, current_y)
          pdf.cell(
              60,
              4,
              clean_ascii(
                  f"Custo Total: R$ {row['Custo Total Estimado (R$)']:,.2f}"
              ),
              0,
              0,
          )
          pdf.set_xy(col2_x, current_y)
          pdf.set_font("Arial", "B", 8)
          pdf.set_text_color(30, 130, 50)
          pdf.cell(
              60,
              4,
              clean_ascii(
                  f"Lucro Líquido: R$ {row['Lucro Líquido Real (R$)']:,.2f}"
              ),
              0,
              0,
          )
          pdf.set_font("Arial", "", 8)
          pdf.set_text_color(50, 60, 75)
          pdf.set_xy(col3_x, current_y)
          pdf.cell(
              60,
              4,
              clean_ascii(
                  f"Pagamento: {row.get('Condições de Pagamento', 'À vista')}"
              ),
              0,
              1,
          )

          current_y += 5
          pdf.set_xy(col1_x, current_y)
          endereco_completo = clean_ascii(
              f"Endereço: {row['Endereço']} - {row['Cidade Imóvel']}/{row['Estado Imóvel']}"
          )
          pdf.cell(180, 4, endereco_completo[:105], 0, 1)

          current_y += 4.5
          pdf.set_xy(col1_x, current_y)
          pdf.set_font("Arial", "U", 7.5)
          pdf.set_text_color(0, 82, 204)
          link_anuncio = str(row["Link do Imóvel"])
          pdf.cell(
              180,
              4,
              clean_ascii(f"Acesse o Link: {link_anuncio}"),
              0,
              1,
              link=(
                  link_anuncio if link_anuncio.startswith("http") else None
              ),
          )

          pdf.ln(6)

    out = pdf.output()
    if isinstance(out, str):
      return out.encode("latin-1")
    return bytes(out)

  def gerar_excel_profissional(df_input):
    output = io.BytesIO()
    df_export = df_input.copy()

    if "Link do Imóvel" in df_export.columns:
      df_export["Link do Imóvel"] = df_export["Link do Imóvel"].apply(
          lambda x: f'=HYPERLINK("{x}", "🔗 Ver Anúncio")'
          if pd.notnull(x) and str(x).startswith("http")
          else x
      )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
      df_export.to_excel(writer, index=False, sheet_name="Oportunidades")
      ws = writer.sheets["Oportunidades"]

      ws.views.sheetView[0].showGridLines = True
      ws.freeze_panes = "A2"
      ws.auto_filter.ref = ws.dimensions

      font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
      fill_header = PatternFill(
          start_color="0052CC", end_color="0052CC", fill_type="solid"
      )
      font_link = Font(
          name="Calibri", size=11, color="0066CC", underline="single"
      )

      align_center = Alignment(horizontal="center", vertical="center")
      align_right = Alignment(horizontal="right", vertical="center")
      align_left = Alignment(horizontal="left", vertical="center")

      col_names = [cell.value for cell in ws[1]]

      for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

      for col_idx, col_name in enumerate(col_names, start=1):
        col_letter = get_column_letter(col_idx)
        cells = ws[col_letter][1:]

        if (
            "Preço" in str(col_name)
            or "Avaliação" in str(col_name)
            or "Lucro" in str(col_name)
            or "Custo" in str(col_name)
        ):
          for c in cells:
            c.number_format = "R$ #,##0.00"
            c.alignment = align_right
        elif "Desconto" in str(col_name):
          for c in cells:
            c.number_format = '0.00"%"'
            c.alignment = align_right
        elif col_name in ["ID Investidor", "Estado Imóvel"]:
          for c in cells:
            c.alignment = align_center
        elif "Link" in str(col_name):
          for c in cells:
            c.font = font_link
            c.alignment = align_center
        else:
          for c in cells:
            c.alignment = align_left

      col_widths = {
          "ID Investidor": 14,
          "Nome do Investidor": 25,
          "Cidades Solicitadas": 25,
          "Faixa Solicitada": 22,
          "Título do Imóvel": 40,
          "Cidade Imóvel": 20,
          "Estado Imóvel": 14,
          "Tipo de Bem": 16,
          "Preço do Leilão (R$)": 22,
          "Valor de Avaliação (R$)": 24,
          "Desconto (%)": 15,
          "Custo Total Estimado (R$)": 24,
          "Lucro Líquido Real (R$)": 22,
          "Endereço": 45,
          "Link do Imóvel": 18,
      }

      for col_idx, col_name in enumerate(col_names, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths.get(col_name, 20)

    return output.getvalue()

  # CABEÇALHO PRINCIPAL
  col_head1, col_head2 = st.columns([4, 1])
  with col_head1:
    st.title("🎯 Cruzador Automático de Leilões & Inteligência")
    st.markdown(
        "Cruzamento inteligente entre o perfil dos investidores e as"
        " oportunidades em leilão."
    )

  with col_head2:
    st.write(" ")
    with st.popover("❓ Como Usar"):
      st.markdown("""
            ### 📖 Passo a Passo Simples
            1. **Carregue as planilhas** abaixo.
            2. Clique em **🚀 Processar Oportunidades**.
            3. Refine por **Preço ou Tipo**.
            4. Navegue pelas **abas do dashboard**.
            """)

  st.divider()

  # CENTRAL DE UPLOAD NA TELA PRINCIPAL
  with st.container():
    st.subheader("📂 Central de Envio e Parâmetros")
    up_col1, up_col2, up_col3 = st.columns([1.5, 1.5, 1])

    with up_col1:
      file_leiloes = st.file_uploader(
          "1️⃣ Base de Leilões (.xlsx)", type=["xlsx", "xls"]
      )
    with up_col2:
      file_investidores = st.file_uploader(
          "2️⃣ Base de Investidores (.xlsx)", type=["xlsx", "xls"]
      )
    with up_col3:
      st.write("⚙️ **Custos Ocultos**")
      taxa_leiloeiro = (
          st.number_input("Comissão Leiloeiro (%)", 0.0, 10.0, 5.0, 0.5) / 100.0
      )
      taxa_itbi = (
          st.number_input("ITBI / Registro (%)", 0.0, 10.0, 3.0, 0.5) / 100.0
      )

    st.write(" ")
    executar = st.button(
        "🚀 Processar Oportunidades e Cruzar Dados", type="primary"
    )

  st.divider()

  # PROCESSAMENTO DE DADOS
  if file_leiloes and file_investidores and executar:
    with st.spinner("Analisando critérios e cruzando bases de dados..."):
      try:
        df_leiloes = pd.read_excel(file_leiloes)
        df_investidores = pd.read_excel(file_investidores)

        df_leiloes["norm_cidade"] = df_leiloes["Cidade"].apply(normalize)
        df_leiloes["norm_estado"] = df_leiloes["Estado"].apply(normalize)
        df_leiloes["norm_tipo"] = df_leiloes["Tipo de Bem"].apply(normalize)

        df_leiloes["preco_effective"] = df_leiloes.apply(
            lambda r: (
                r["2º Leilão (Preço)"]
                if pd.notnull(r["2º Leilão (Preço)"])
                and r["2º Leilão (Preço)"] > 0
                else r["1º Leilão (Preço)"]
            ),
            axis=1,
        )

        df_leiloes["desconto_%"] = np.where(
            (pd.notnull(df_leiloes["Valor de Avaliação do Leiloeiro"]))
            & (df_leiloes["Valor de Avaliação do Leiloeiro"] > 0)
            & (pd.notnull(df_leiloes["preco_effective"]))
            & (df_leiloes["preco_effective"] > 0),
            ((
                df_leiloes["Valor de Avaliação do Leiloeiro"]
                - df_leiloes["preco_effective"]
            )
            / df_leiloes["Valor de Avaliação do Leiloeiro"])
            * 100,
            0,
        )

        cols = list(df_investidores.columns)
        resultados = []

        for idx, row in df_investidores.iterrows():
          nome = str(row["Nome Completo"]).strip()
          cidades_input = str(row[cols[10]]).strip()
          tipos_input = str(row[cols[7]]).strip()
          valor_input = str(row[cols[8]]).strip()
          cons = str(row[cols[15]]).strip()

          norm_cid = normalize(cidades_input)
          norm_cons = normalize(cons)
          norm_val = normalize(valor_input)

          sub = df_leiloes.copy()

          target_cidades = []
          target_estados = []

          if any(
              w in norm_cid or w in norm_cons
              for w in [
                  "rio preto",
                  "sao jose do rio preto",
                  "sao jose do preto",
                  "sjrp",
              ]
          ):
            target_cidades.append("sao jose do rio preto")
          if "bady" in norm_cid or "bady" in norm_cons:
            target_cidades.append("bady bassitt")
          if "mirassol" in norm_cid or "mirassol" in norm_cons:
            target_cidades.append("mirassol")
          if "santo andre" in norm_cid or "santo andre" in norm_cons:
            target_cidades.append("santo andre")
          if "sao bernardo" in norm_cid or "sao bernardo" in norm_cons:
            target_cidades.append("sao bernardo do campo")
          if "sao caetano" in norm_cid or "sao caetano" in norm_cons:
            target_cidades.append("sao caetano do sul")
          if "abc" in norm_cid or "abc" in norm_cons:
            target_cidades.extend(
                ["santo andre", "sao bernardo do campo", "sao caetano do sul"]
            )
          if "sao paulo" in norm_cid or "sao paulo" in norm_cons:
            if "grande sao paulo" in norm_cid or "grande sao paulo" in norm_cons:
              target_cidades.extend([
                  "sao paulo",
                  "santo andre",
                  "sao bernardo do campo",
                  "sao caetano do sul",
              ])
            elif "estado de sp" in norm_cid or "estado de sp" in norm_cons:
              target_estados.append("sao paulo")
            else:
              target_cidades.append("sao paulo")
          if "sorocaba" in norm_cid or "sorocaba" in norm_cons:
            target_cidades.append("sorocaba")
          if "tatui" in norm_cid or "tatui" in norm_cons:
            target_cidades.append("tatui")
          if "boituva" in norm_cid or "boituva" in norm_cons:
            target_cidades.append("boituva")
          if "rio de janeiro" in norm_cid or "rio de janeiro" in norm_cons:
            target_cidades.append("rio de janeiro")
          if "goiania" in norm_cid:
            target_cidades.append("goiania")
          if "brasilia" in norm_cid:
            target_cidades.append("brasilia")
          if "anapolis" in norm_cid:
            target_cidades.append("anapolis")
          if "cidades polo" in norm_cid or "expansao" in norm_cid:
            target_estados.append("sao paulo")

          target_cidades = list(set(target_cidades))
          target_estados = list(set(target_estados))

          if target_cidades:
            sub = sub[sub["norm_cidade"].isin(target_cidades)]
          elif target_estados:
            sub = sub[sub["norm_estado"].isin(target_estados)]

          allowed_types = parse_types(tipos_input)
          if "terrenos" in norm_cons or "lotes" in norm_cons:
            if "Terreno" not in allowed_types:
              allowed_types.append("Terreno")

          if allowed_types:
            sub = sub[sub["Tipo de Bem"].isin(allowed_types)]

          min_v, max_v = parse_budget(valor_input)
          if any(
              p in norm_cons or p in norm_val
              for p in [
                  "qualquer valor",
                  "valores menores",
                  "100 a 500",
                  "todos os valores",
                  "qualquer",
              ]
          ):
            min_v, max_v = 0, 999999999

          if min_v > 0 and max_v < 999999999:
            sub = sub[
                (sub["preco_effective"] >= min_v)
                & (sub["preco_effective"] <= max_v)
            ]
          elif min_v > 0:
            sub = sub[sub["preco_effective"] >= min_v]
          elif max_v < 999999999:
            sub = sub[sub["preco_effective"] <= max_v]

          sub_sorted = sub.sort_values(
              by=["desconto_%", "preco_effective"], ascending=[False, True]
          )

          for _, imovel in sub_sorted.iterrows():
            preco = imovel["preco_effective"]
            avaliac = imovel["Valor de Avaliação do Leiloeiro"]

            custos_adicionais = (preco * taxa_leiloeiro) + (preco * taxa_itbi)
            custo_total = preco + custos_adicionais
            lucro_liquido = (
                avaliac - custo_total
                if pd.notnull(avaliac) and pd.notnull(preco)
                else 0
            )

            resultados.append({
                "ID Investidor": idx + 1,
                "Nome do Investidor": nome,
                "Cidades Solicitadas": cidades_input,
                "Faixa Solicitada": valor_input,
                "Título do Imóvel": imovel["Título"],
                "Cidade Imóvel": imovel["Cidade"],
                "Estado Imóvel": imovel["Estado"],
                "Tipo de Bem": imovel["Tipo de Bem"],
                "Preço do Leilão (R$)": preco,
                "Valor de Avaliação (R$)": avaliac,
                "Desconto (%)": round(imovel["desconto_%"], 2),
                "Custo Total Estimado (R$)": round(custo_total, 2),
                "Lucro Líquido Real (R$)": round(lucro_liquido, 2),
                "Endereço": imovel["Endereço"],
                "Link do Imóvel": imovel["Link"],
            })

        st.session_state["df_final"] = pd.DataFrame(resultados)
        st.session_state["imoveis_selecionados"] = []
        st.toast("✅ Processamento Enterprise concluído!", icon="🎉")

      except Exception as e:
        st.error(f"Erro ao processar as planilhas: {e}")

  # EXIBIÇÃO DO DASHBOARD E ABAS
  if "df_final" in st.session_state and not st.session_state["df_final"].empty:
    df_base = st.session_state["df_final"]

    with st.expander("🔍 **Filtros Avançados de Refinamento**", expanded=True):
      f_col1, f_col2, f_col3 = st.columns([2, 2, 2])

      with f_col1:
        max_p = (
            float(df_base["Preço do Leilão (R$)"].max())
            if not df_base.empty
            else 5000000.0
        )
        limite_slider = max(max_p, 100000000.0)
        faixa_preco = st.slider(
            "Faixa de Preço do Leilão (R$)",
            min_value=0.0,
            max_value=limite_slider,
            value=(0.0, limite_slider),
            step=50000.0,
            format="R$ %,.0f",
        )

      with f_col2:
        tipos_disponiveis = sorted(df_base["Tipo de Bem"].unique().tolist())
        tipos_selecionados = st.multiselect(
            "Tipo de Bem",
            options=tipos_disponiveis,
            default=tipos_disponiveis,
        )

      with f_col3:
        busca_texto = st.text_input(
            "Buscar Palavra-chave",
            value="",
            placeholder="Ex: Rio Preto, Simonetti, Casa...",
        )

    df_filtered = df_base[
        (df_base["Preço do Leilão (R$)"] >= faixa_preco[0])
        & (df_base["Preço do Leilão (R$)"] <= faixa_preco[1])
        & (df_base["Tipo de Bem"].isin(tipos_selecionados))
    ]

    if busca_texto.strip():
      termo = normalize(busca_texto)
      df_filtered = df_filtered[
          df_filtered["Cidade Imóvel"].apply(normalize).str.contains(termo)
          | df_filtered["Nome do Investidor"]
          .apply(normalize)
          .str.contains(termo)
          | df_filtered["Título do Imóvel"].apply(normalize).str.contains(termo)
      ]

    num_sel = (
        len(st.session_state["imoveis_selecionados"])
        if "imoveis_selecionados" in st.session_state
        else 0
    )

    tab_labels = (
        [
            "📊 Visão Geral",
            "📋 Tabela de Oportunidades",
            f"⭐ Selecionados ({num_sel})",
            "👤 Vitrine / Cards por Investidor",
        ]
        if is_tester
        else [
            "📊 Visão Geral",
            "📋 Tabela de Oportunidades & Download",
            f"⭐ Selecionados ({num_sel})",
            "👤 Vitrine / Cards por Investidor (PDF / WhatsApp)",
        ]
    )

    tabs = st.tabs(tab_labels)
    tab1, tab2, tab3, tab4 = tabs[0], tabs[1], tabs[2], tabs[3]

    with tab1:
      st.write(" ")
      kpi1, kpi2, kpi3, kpi4 = st.columns(4)
      kpi1.metric("🏢 Total Imóveis", f"{len(df_filtered):,}")
      kpi2.metric(
          "👥 Investidores",
          f"{df_filtered['Nome do Investidor'].nunique()}"
          if not df_filtered.empty
          else "0",
      )
      kpi3.metric(
          "🔥 Maior Desconto",
          f"{df_filtered['Desconto (%)'].max():.1f}%"
          if not df_filtered.empty
          else "0%",
      )
      total_lucro_liq = (
          df_filtered["Lucro Líquido Real (R$)"].sum()
          if not df_filtered.empty
          else 0
      )
      kpi4.metric("💰 Lucro Líquido Acumulado", f"R$ {total_lucro_liq:,.0f}")

      st.markdown("---")

      if not df_filtered.empty:
        g_col1, g_col2 = st.columns(2)

        with g_col1:
          st.subheader("🍩 Distribuição por Tipo de Imóvel")
          fig_pie = px.pie(
              df_filtered,
              names="Tipo de Bem",
              hole=0.4,
              color_discrete_sequence=px.colors.qualitative.Bold,
          )
          fig_pie.update_layout(margin=dict(t=20, b=20, l=10, r=10), height=320)
          st.plotly_chart(fig_pie, use_container_width=True)

        with g_col2:
          st.subheader("📍 Top 5 Cidades com Mais Oportunidades")
          top_cidades = (
              df_filtered["Cidade Imóvel"]
              .value_counts()
              .head(5)
              .reset_index()
          )
          top_cidades.columns = ["Cidade", "Total"]
          fig_bar = px.bar(
              top_cidades,
              x="Total",
              y="Cidade",
              orientation="h",
              text="Total",
              color_discrete_sequence=["#0052CC"],
          )
          fig_bar.update_layout(
              margin=dict(t=20, b=20, l=10, r=10),
              height=320,
              yaxis={"categoryorder": "total ascending"},
          )
          st.plotly_chart(fig_bar, use_container_width=True)

    with tab2:
      st.write(" ")
      if not df_filtered.empty:
        d_col1, d_col2 = st.columns([3, 1])
        with d_col1:
          st.subheader("📋 Tabela Consolidada de Oportunidades")
        with d_col2:
          if not is_tester:
            excel_bytes = gerar_excel_profissional(df_filtered)
            st.download_button(
                label="📥 Baixar Excel Formatado",
                data=excel_bytes,
                file_name="cruzamento_leiloes_investidores.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
                use_container_width=True,
            )
          else:
            st.info("🔒 Download em Excel restrito no modo de teste.")

        st.dataframe(
            df_filtered,
            use_container_width=True,
            height=500,
            column_config={
                "Link do Imóvel": st.column_config.LinkColumn(
                    "Anúncio Oficial", display_text="🔗 Ver Imóvel"
                ),
                "Preço do Leilão (R$)": st.column_config.NumberColumn(
                    "Preço Leilão", format="R$ %,.2f"
                ),
                "Valor de Avaliação (R$)": st.column_config.NumberColumn(
                    "Valor Avaliação", format="R$ %,.2f"
                ),
                "Custo Total Estimado (R$)": st.column_config.NumberColumn(
                    "Custo Total", format="R$ %,.2f"
                ),
                "Lucro Líquido Real (R$)": st.column_config.NumberColumn(
                    "Lucro Líquido Real", format="R$ %,.2f"
                ),
                "Desconto (%)": st.column_config.ProgressColumn(
                    "Desconto (%)", format="%.1f%%", min_value=0, max_value=100
                ),
            },
        )
      else:
        st.warning(
            "Nenhum imóvel encontrado com os filtros selecionados acima."
        )

    with tab3:
      st.write(" ")
      s_col1, s_col2, s_col3 = st.columns([2, 1.5, 1.5])
      with s_col1:
        st.subheader("⭐ Seus Imóveis Escolhidos / Selecionados")
        st.markdown(
            "Aqui estão reunidos todos os imóveis que você marcou na vitrine."
        )
      with s_col2:
        if (
            "imoveis_selecionados" in st.session_state
            and st.session_state["imoveis_selecionados"]
            and not is_tester
        ):
          st.write(" ")
          df_sel_exp = pd.DataFrame(st.session_state["imoveis_selecionados"])
          excel_sel_bytes = gerar_excel_profissional(df_sel_exp)
          st.download_button(
              label="📥 Baixar Excel Selecionados",
              data=excel_sel_bytes,
              file_name="imoveis_selecionados.xlsx",
              mime=(
                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
              ),
              use_container_width=True,
          )
      with s_col3:
        if (
            "imoveis_selecionados" in st.session_state
            and st.session_state["imoveis_selecionados"]
            and not is_tester
        ):
          st.write(" ")
          df_sel_exp = pd.DataFrame(st.session_state["imoveis_selecionados"])
          pdf_sel_bytes = gerar_pdf_informativo(
              "Imóveis Selecionados", df_sel_exp
          )
          st.download_button(
              label="📄 Baixar PDF Informativo",
              data=pdf_sel_bytes,
              file_name="informativo_selecionados.pdf",
              mime="application/pdf",
              use_container_width=True,
          )

      if (
          "imoveis_selecionados" in st.session_state
          and st.session_state["imoveis_selecionados"]
      ):
        df_sel = pd.DataFrame(st.session_state["imoveis_selecionados"])

        if st.button("🗑️ Limpar Todos os Selecionados"):
          st.session_state["imoveis_selecionados"] = []
          st.rerun()

        st.write(" ")

        cols_sel = st.columns(2)
        for idx, row in df_sel.iterrows():
          col_target = cols_sel[idx % 2]
          link_url = (
              row["Link do Imóvel"]
              if str(row["Link do Imóvel"]).startswith("http")
              else "#"
          )

          with col_target:
            st.markdown(
                f"""
                        <div class="property-card">
                            <span class="badge-type">🏠 {row['Tipo de Bem']}</span>
                            <h4 style="margin-top: 8px; margin-bottom: 4px; color: #1E293B;">{row['Título do Imóvel']}</h4>
                            <p style="color: #64748B; font-size: 0.9rem; margin-bottom: 8px;">📍 {row['Cidade Imóvel']} - {row['Estado Imóvel']} | Pretendente: <b>{row['Nome do Investidor']}</b></p>
                            <div style="margin-bottom: 8px;">
                                <span class="price-main">R$ {row['Preço do Leilão (R$)']:,.2f}</span> <span class="price-old">(Avaliação: R$ {row['Valor de Avaliação (R$)']:,.2f})</span><br>
                                <span class="price-profit">💰 Lucro Líquido Real: R$ {row['Lucro Líquido Real (R$)']:,.2f}</span>
                            </div>
                            <p style="font-size: 0.85rem; color: #475569; margin-bottom: 12px;"><b>Endereço:</b> {row['Endereço']}</p>
                            <a href="{link_url}" target="_blank" style="text-decoration: none;">
                                <button style="background-color: #0052CC; color: white; border: none; padding: 8px; border-radius: 6px; font-weight: bold; cursor: pointer; width: 100%;">
                                    🔗 Ver Anúncio Oficial
                                </button>
                            </a>
                        </div>
                        """,
                unsafe_allow_html=True,
            )
      else:
        st.info(
            "💡 Nenhum imóvel foi selecionado ainda. Vá até a aba **👤 Vitrine /"
            " Cards por Investidor** e clique no botão **⭐ Escolher Imóvel**"
            " nos cards que desejar."
        )

    with tab4:
      st.write(" ")
      if not df_filtered.empty:
        investidores_lista = sorted(
            df_filtered["Nome do Investidor"].unique().tolist()
        )

        c_sel1, c_sel2, c_sel3 = st.columns([2, 1, 1])
        with c_sel1:
          investidor_sel = st.selectbox(
              "👤 Selecione o Investidor:",
              options=investidores_lista,
              key="select_inv_vitrine",
          )
        with c_sel2:
          st.write(" ")
          df_inv_curr = df_filtered[
              df_filtered["Nome do Investidor"] == investidor_sel
          ]
          if not df_inv_curr.empty:
            if not is_tester:
              pdf_bytes = gerar_pdf_informativo(investidor_sel, df_inv_curr)
              st.download_button(
                  label="📄 PDF Informativo 1",
                  data=pdf_bytes,
                  file_name=(
                      f"informativo_1_{normalize(investidor_sel).replace(' ', '_')}.pdf"
                  ),
                  mime="application/pdf",
                  use_container_width=True,
              )
            else:
              st.info("🔒 PDF restrito.")
        with c_sel3:
          st.write(" ")
          if not df_inv_curr.empty:
            if not is_tester:
              pdf_bytes2 = gerar_pdf_informativo(investidor_sel, df_inv_curr)
              st.download_button(
                  label="📄 PDF Informativo 2",
                  data=pdf_bytes2,
                  file_name=(
                      f"informativo_2_{normalize(investidor_sel).replace(' ', '_')}.pdf"
                  ),
                  mime="application/pdf",
                  use_container_width=True,
              )
            else:
              st.info("🔒 PDF restrito.")

        df_inv = df_filtered[
            df_filtered["Nome do Investidor"] == investidor_sel
        ]

        st.markdown(f"### 🎯 Vitrine Exclusiva: **{investidor_sel}**")
        st.caption(
            f"Preferências solicitadas: {df_inv['Cidades Solicitadas'].iloc[0]}"
            f" | Faixa: {df_inv['Faixa Solicitada'].iloc[0]}"
        )

        st.write(" ")

        itens_por_pagina = 10
        total_imoveis = len(df_inv)
        total_pages = (
            (total_imoveis + itens_por_pagina - 1) // itens_por_pagina
            if total_imoveis > 0
            else 1
        )

        if total_pages > 1:
          pagina_atual = st.number_input(
              "📄 Página de Cards",
              min_value=1,
              max_value=total_pages,
              step=1,
              key="pag_cards",
          )
        else:
          pagina_atual = 1

        start_idx = (pagina_atual - 1) * itens_por_pagina
        end_idx = start_idx + itens_por_pagina
        df_paginado = df_inv.iloc[start_idx:end_idx]

        st.caption(
            f"Mostrando imóveis {start_idx + 1} a"
            f" {min(end_idx, total_imoveis)} de {total_imoveis} para este"
            " investidor."
        )

        @st.fragment
        def renderizar_vitrine_v25(df_cards):
          cols_cards = st.columns(2)
          for idx, (_, row) in enumerate(df_cards.iterrows()):
            col_target = cols_cards[idx % 2]

            badge_html = f'<span class="badge-type">🏠 {row["Tipo de Bem"]}</span>'

            # Cálculo individual dos custos ocultos para exibição nos cards
            preco_card = row["Preço do Leilão (R$)"]
            val_comissao_leiloeiro = preco_card * taxa_leiloeiro
            val_itbi_cartorio = preco_card * taxa_itbi

            link_url = (
                row["Link do Imóvel"]
                if str(row["Link do Imóvel"]).startswith("http")
                else "#"
            )

            ja_selecionado = any(
                item["Título do Imóvel"] == row["Título do Imóvel"]
                and item["Nome do Investidor"] == row["Nome do Investidor"]
                for item in st.session_state["imoveis_selecionados"]
            )

            with col_target:
              # Container principal do card com o botão de favoritar (estrela) no topo direito
              with st.container():
                st.markdown(
                    f"""
                    <div class="property-card">
                        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 4px;">
                            <div>{badge_html}</div>
                        </div>
                        <h4 style="margin-top: 4px; margin-bottom: 4px; color: #1E293B;">{row['Título do Imóvel']}</h4>
                        <p style="color: #64748B; font-size: 0.9rem; margin-bottom: 8px;">📍 {row['Cidade Imóvel']} - {row['Estado Imóvel']}</p>
                        <div style="margin-bottom: 8px;">
                            <span class="price-main">R$ {row['Preço do Leilão (R$)']:,.2f}</span> <span class="price-old">(Avaliação: R$ {row['Valor de Avaliação (R$)']:,.2f})</span><br>
                            <span class="price-profit">💰 Lucro Líquido Real: R$ {row['Lucro Líquido Real (R$)']:,.2f}</span><br>
                            <span class="price-costs">• Comissão do Leiloeiro: R$ {val_comissao_leiloeiro:,.2f}</span><br>
                            <span class="price-costs">• ITBI e Cartório: R$ {val_itbi_cartorio:,.2f}</span><br>
                            <span class="price-costs">🛠️ Custo Total Estimado: R$ {row['Custo Total Estimado (R$)']:,.2f}</span>
                        </div>
                        <p style="font-size: 0.85rem; color: #475569; margin-bottom: 12px;"><b>Endereço:</b> {row['Endereço']}</p>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

              # Botões de Estrela (Favoritar) e Link de Anúncio integrados lado a lado
              sub_col1, sub_col2 = st.columns([1, 2])
              with sub_col1:
                label_estrela = "⭐ Favoritado" if ja_selecionado else "☆ Favoritar"
                if st.button(
                    label_estrela,
                    key=f"btn_estrela_{idx}_{row['Título do Imóvel']}",
                    use_container_width=True,
                ):
                  if ja_selecionado:
                    st.session_state["imoveis_selecionados"] = [
                        item
                        for item in st.session_state["imoveis_selecionados"]
                        if not (
                            item["Título do Imóvel"] == row["Título do Imóvel"]
                            and item["Nome do Investidor"]
                            == row["Nome do Investidor"]
                        )
                    ]
                  else:
                    st.session_state["imoveis_selecionados"].append(
                        row.to_dict()
                    )
                  st.rerun()

              with sub_col2:
                st.markdown(
                    f"""
                    <a href="{link_url}" target="_blank" style="text-decoration: none;">
                        <button style="background-color: #0052CC; color: white; border: none; padding: 10px; border-radius: 6px; font-weight: bold; cursor: pointer; width: 100%; font-size: 0.9rem;">
                            🔗 Ver Anúncio Oficial
                        </button>
                    </a>
                    """,
                    unsafe_allow_html=True,
                )

              st.write("---")

        renderizar_vitrine_v25(df_paginado)

  elif "df_final" not in st.session_state:
    st.info(
        "💡 **Para iniciar:** Faça o upload das duas planilhas e ajuste os"
        " custos na **Central de Envio** acima, depois clique em **🚀 Processar"
        " Oportunidades**."
    )
